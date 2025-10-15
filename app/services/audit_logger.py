from pathlib import Path
from datetime import datetime
from typing import Any, Dict , cast

import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

from openpyxl.worksheet.worksheet import Worksheet

# === Rutas / Salida ===
OUTPUTS_DIR = Path(__file__).resolve().parents[3] / "outputs"
OUTPUTS_DIR.mkdir(parents=True, exist_ok=True)
XLSX_PATH = OUTPUTS_DIR / "resultados_auditoria.xlsx"

# === Esquema fijo de columnas (en orden) ===
COLUMNS = [
    "timestamp",
    "archivo",
    "ruta",
    # Presencia (flags)
    "cedula",
    "fecha",
    "medicamento",
    "cantidad",
    "firma",
    "metodo_firma",
    # Extraído (valores OCR/parse)
    "doc_extraido",
    "fecha_extraida",
    "medicamento_extraido",
    "cantidad_extraida",
    # Comparación (booleans SI/NO)
    "doc_ok",
    "fecha_ok",
    "med_ok",
    "cant_ok",
    # Esperado (desde tabla de referencia)
    "doc_esperado",
    "fecha_esperada",
    "medicamento_esperado",
    "cantidad_esperada",
    # Resumen
    "faltantes",
    "observaciones",
]

# === Helpers visuales ===
def _mark(v: Any) -> str:
    return "✅" if bool(v) else "❌"

def _si_no(v: Any) -> str:
    if v is True:
        return "SI"
    if v is False:
        return "NO"
    return ""  # vacío cuando no hay referencia

def _safe_get(d: Dict[str, Any], key: str, default: str = "") -> str:
    v = d.get(key, default) if isinstance(d, dict) else default
    if v is None:
        return ""
    return str(v)

def _ensure_schema(df: pd.DataFrame) -> pd.DataFrame:
    """Alinea DataFrame al esquema COLUMNS (agrega faltantes y reordena)."""
    for col in COLUMNS:
        if col not in df.columns:
            df[col] = ""
    return df[COLUMNS]

def _auto_format_excel(path: Path) -> None:
    """Aplica formato seguro (encabezados, filtros, anchos, wrap...) al Excel."""
    if not path.exists():
        return

    try:
        wb = load_workbook(path)

        # Asegurar al menos una hoja y tomarla por nombre
        if not wb.sheetnames:
            ws = wb.create_sheet(title="Resultados")
        else:
            # Toma explícitamente la primera hoja por nombre y tipa Worksheet
            ws = cast(Worksheet, wb[wb.sheetnames[0]])

        # Si no hay datos, no hay nada que formatear
        if ws.max_row < 1 or ws.max_column < 1:
            wb.save(path)
            return

        # 1) Congelar encabezado
        ws.freeze_panes = "A2"

        # 2) Negrita en encabezados + Autofiltro
        header_font = Font(bold=True)
        max_col = ws.max_column
        max_row = ws.max_row
        ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"
        for col in range(1, max_col + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font

        # 3) Anchos de columna sugeridos (ajusta si agregas/quitas columnas)
        widths = {
            "A": 19,  # timestamp
            "B": 28,  # archivo
            "C": 60,  # ruta
            "D": 10,  # cedula (flag)
            "E": 10,  # fecha (flag)
            "F": 14,  # medicamento (flag)
            "G": 12,  # cantidad (flag)
            "H": 10,  # firma (flag)
            "I": 14,  # metodo_firma
            "J": 18,  # doc_extraido
            "K": 16,  # fecha_extraida
            "L": 26,  # medicamento_extraido
            "M": 16,  # cantidad_extraida
            "N": 10,  # doc_ok
            "O": 10,  # fecha_ok
            "P": 10,  # med_ok
            "Q": 10,  # cant_ok
            "R": 18,  # doc_esperado
            "S": 16,  # fecha_esperada
            "T": 26,  # medicamento_esperado
            "U": 16,  # cantidad_esperada
            "V": 28,  # faltantes
            "W": 48,  # observaciones
        }
        for col_letter, w in widths.items():
            ws.column_dimensions[col_letter].width = w

        # 4) Wrap en columnas largas
        wrap_align = Alignment(wrap_text=True, vertical="top")
        cols_to_wrap_names = ["ruta", "medicamento_extraido", "medicamento_esperado", "faltantes", "observaciones"]
        for name in cols_to_wrap_names:
            if name in COLUMNS:
                col_idx = COLUMNS.index(name) + 1  # 1-based idx
                for r in range(2, ws.max_row + 1):
                    ws.cell(row=r, column=col_idx).alignment = wrap_align

        wb.save(path)

    except Exception as e:
        # En producción podrías loggear esto
        # print(f"[WARN] No se pudo formatear Excel: {e}")
        pass



def log_result(payload: Dict[str, Any]) -> str:
    """
    payload:
      filename, path, result:{
        firma(bool), firma_method(str|None),
        cedula, medicamento, fecha, cantidad (bools),
        faltantes(list[str]|str),
        observaciones(str),
        extraido: { documento, fecha_pedido, medicamento, cantidad },
        comparacion: {
          documento_ok(bool|None),
          fecha_ok(bool|None),
          medicamento_ok(bool|None),
          cantidad_ok(bool|None),
          esperado: { documento, fecha_pedido, medicamento, cantidad } | ausente
        }
      }
    """
    r = payload.get("result", {}) or {}
    extraido = r.get("extraido", {}) if isinstance(r, dict) else {}
    comp = r.get("comparacion", {}) if isinstance(r, dict) else {}
    esperado = comp.get("esperado", {}) if isinstance(comp, dict) else {}

    # Normaliza faltantes a texto legible (sin duplicados, orden estable)
    faltantes_list = r.get("faltantes", [])
    if isinstance(faltantes_list, list):
        faltantes_txt = ", ".join(dict.fromkeys(faltantes_list))
    else:
        faltantes_txt = str(faltantes_list or "")

    # Observaciones limpias
    obs = (r.get("observaciones", "") or "").replace("\t", " ").replace("\n", " ").strip()[:500]

    row = {
        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "archivo": payload.get("filename", ""),
        "ruta": payload.get("path", ""),

        # Presencia (flags)
        "cedula": _mark(r.get("cedula", False)),
        "fecha": _mark(r.get("fecha", False)),
        "medicamento": _mark(r.get("medicamento", False)),
        "cantidad": _mark(r.get("cantidad", False)),
        "firma": _mark(r.get("firma", False)),
        "metodo_firma": _safe_get(r, "firma_method"),

        # Extraído
        "doc_extraido": _safe_get(extraido, "documento"),
        "fecha_extraida": _safe_get(extraido, "fecha_pedido"),
        "medicamento_extraido": _safe_get(extraido, "medicamento"),
        "cantidad_extraida": _safe_get(extraido, "cantidad"),

        # Comparación (SI/NO o vacío si no hay referencia)
        "doc_ok": _si_no(comp.get("documento_ok") if isinstance(comp, dict) else None),
        "fecha_ok": _si_no(comp.get("fecha_ok") if isinstance(comp, dict) else None),
        "med_ok": _si_no(comp.get("medicamento_ok") if isinstance(comp, dict) else None),
        "cant_ok": _si_no(comp.get("cantidad_ok") if isinstance(comp, dict) else None),

        # Esperado
        "doc_esperado": _safe_get(esperado, "documento"),
        "fecha_esperada": _safe_get(esperado, "fecha_pedido"),
        "medicamento_esperado": _safe_get(esperado, "medicamento"),
        "cantidad_esperada": _safe_get(esperado, "cantidad"),

        # Resumen
        "faltantes": faltantes_txt,
        "observaciones": obs,
    }

    df_new = pd.DataFrame([row], columns=COLUMNS)

    if XLSX_PATH.exists():
        existing = pd.read_excel(XLSX_PATH)
        existing = _ensure_schema(existing)

        # Evitar duplicados: elimina filas previas del mismo archivo
        existing = existing[existing["archivo"] != row["archivo"]]

        df_out = pd.concat([existing, df_new], ignore_index=True)
    else:
        # Crear un workbook vacío si no existe, para evitar problemas de estilos en la primera pasada
        df_out = df_new

    # Orden consistente: por 'archivo' (a-z) y timestamp asc
    df_out["_archivo_sort"] = df_out["archivo"].astype(str).str.casefold()
    try:
        ts = pd.to_datetime(df_out["timestamp"], errors="coerce", format="%Y-%m-%d %H:%M:%S")
    except Exception:
        ts = pd.to_datetime(df_out["timestamp"], errors="coerce")
    df_out["_ts_sort"] = ts

    df_out = df_out.sort_values(by=["_archivo_sort", "_ts_sort"], ascending=[True, True], ignore_index=True)
    df_out = df_out.drop(columns=["_archivo_sort", "_ts_sort"], errors="ignore")
    df_out = _ensure_schema(df_out)

    # Guardar y formatear
    df_out.to_excel(XLSX_PATH, index=False)
    _auto_format_excel(XLSX_PATH)
    return str(XLSX_PATH)

